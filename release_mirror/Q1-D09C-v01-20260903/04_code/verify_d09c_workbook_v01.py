from __future__ import annotations

import csv
import hashlib
import json
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(r"C:\range_paper")
WORK = ROOT / "99_tmp" / "d09c_v01"
OUTPUT = WORK / "outputs"
QC = WORK / "qc"
XLSX = OUTPUT / "Q1_D09C_MAINLINE_AUDIT_v01.xlsx"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def data_rows(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.reader(handle)) - 1


checks: list[dict[str, object]] = []


def check(name: str, observed: object, expected: object, passed: bool | None = None) -> None:
    status = bool(observed == expected) if passed is None else bool(passed)
    checks.append({"check": name, "status": "PASS" if status else "FAIL", "observed": observed, "expected": expected})


with zipfile.ZipFile(XLSX) as archive:
    bad_member = archive.testzip()
check("xlsx_container_integrity", bad_member, None)

book = openpyxl.load_workbook(XLSX, read_only=False, data_only=False)
expected_sheets = [
    "Summary",
    "Frame Summary",
    "Frame Audit",
    "EVALID Components",
    "Candidates",
    "Top Partitions",
    "Lineage",
    "Calibration",
    "Input Audit",
    "Table Access",
    "SQL Ledger",
    "QC",
    "Large Ledger Index",
]
check("worksheet_set_exact", sorted(book.sheetnames), sorted(expected_sheets))

sheet_sources = {
    "Frame Summary": "D09C_TEMPORAL_FRAME_SUMMARY_v01.csv",
    "Frame Audit": "D09C_TEMPORAL_FRAME_AUDIT_v01.csv",
    "EVALID Components": "D09C_EVALID_COMPONENT_LEDGER_v01.csv",
    "Candidates": "D09C_WHOLE_PANEL_PARTITION_CANDIDATES_v01.csv",
    "Top Partitions": "D09C_TOP_PARTITION_DIAGNOSTICS_v01.csv",
    "Lineage": "D09C_LINEAGE_CROSSFOLD_AUDIT_v01.csv",
    "Calibration": "D09C_DESIGN_CALIBRATION_AUDIT_v01.csv",
    "Input Audit": "D09C_INPUT_AUDIT_v01.csv",
    "Table Access": "D09C_TABLE_ACCESS_AUDIT_v01.csv",
    "SQL Ledger": "D09C_SQL_LEDGER_v01.csv",
    "QC": "D09C_QC_v01.csv",
}
observed_rows = {sheet: book[sheet].max_row - 1 for sheet in sheet_sources}
expected_rows = {sheet: data_rows(OUTPUT / file_name) for sheet, file_name in sheet_sources.items()}
check("worksheet_data_row_counts_match_csv", observed_rows, expected_rows)

summary = book["Summary"]
check("summary_execution_formula", summary["B3"].value, '=IF(COUNTIF(\'QC\'!$B$2:$B$20,"FAIL")=0,"PASS","FAIL")')
check("summary_feasibility_formula", summary["D3"].value, '=IF(COUNTIF(\'Frame Summary\'!$H$2:$H$3,"FAIL")>0,"FAIL","PASS")')
check("summary_stop_boundaries", [summary["F3"].value, summary["H3"].value], ["NOT SELECTED", "NONE"])

frame = book["Frame Audit"]
components = book["EVALID Components"]
identifier_formulas = {
    "Frame Audit!C2": frame["C2"].value,
    "Frame Audit!G2": frame["G2"].value,
    "EVALID Components!B2": components["B2"].value,
    "EVALID Components!F2": components["F2"].value,
    "EVALID Components!G2": components["G2"].value,
    "EVALID Components!I2": components["I2"].value,
    "EVALID Components!K2": components["K2"].value,
}
expected_identifier_formulas = {
    "Frame Audit!C2": '="01"',
    "Frame Audit!G2": '="012022"',
    "EVALID Components!B2": '="01"',
    "EVALID Components!F2": '="012022"',
    "EVALID Components!G2": '="1263005816290487"',
    "EVALID Components!I2": '="1263004985290487"',
    "EVALID Components!K2": '="1263004974290487"',
}
check("identifier_text_formulas_preserve_leading_zeros_and_64bit_ids", identifier_formulas, expected_identifier_formulas)

large_files = [
    ("D09C_PANEL_SUBPANEL_YEAR_LEDGER_v01.csv", 65233),
    ("D09C_TI_FOLD_WEIGHT_AUDIT_v01.csv", 9472),
    ("D09C_MA_FOLD_WEIGHT_AUDIT_v01.csv", 23680),
]
large_index = book["Large Ledger Index"]
observed_large = []
for row_number, (file_name, expected_count) in enumerate(large_files, start=2):
    file_path = OUTPUT / file_name
    observed_large.append(
        [
            large_index.cell(row_number, 1).value,
            int(large_index.cell(row_number, 2).value),
            int(large_index.cell(row_number, 3).value),
            large_index.cell(row_number, 4).value,
        ]
    )
expected_large = [[name, count, (OUTPUT / name).stat().st_size, sha256(OUTPUT / name)] for name, count in large_files]
check("large_ledger_index_hashes_rows_bytes", observed_large, expected_large)

render_index = json.loads((QC / "D09C_WORKBOOK_RENDER_INDEX_v01.json").read_text(encoding="utf-8"))
render_files = [Path(item["preview_path"]) for item in render_index]
check("all_worksheets_rendered", len(render_index), 13)
check("all_rendered_previews_present_nonempty", [path.exists() and path.stat().st_size > 0 for path in render_files], [True] * 13)

formula_scan = (QC / "D09C_WORKBOOK_FORMULA_ERROR_SCAN_v01.ndjson").read_text(encoding="utf-8").strip()
check("formula_error_scan_zero", formula_scan, '{"kind":"notice","message":"Cell search matched 0 entries."}')

failed = [item["check"] for item in checks if item["status"] != "PASS"]
audit = {
    "audit_id": "D09C_WORKBOOK_INDEPENDENT_AUDIT_v01",
    "date": "2026-09-02",
    "status": "PASS" if not failed else "FAIL",
    "check_count": len(checks),
    "failed": failed,
    "workbook": str(XLSX),
    "workbook_bytes": XLSX.stat().st_size,
    "workbook_sha256": sha256(XLSX),
    "checks": checks,
}
(QC / "D09C_WORKBOOK_INDEPENDENT_AUDIT_v01.json").write_text(json.dumps(audit, indent=2) + "\n", encoding="utf-8")
print(json.dumps({"status": audit["status"], "check_count": len(checks), "failed": failed}, indent=2))
if failed:
    raise SystemExit(1)
